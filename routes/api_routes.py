"""
API Routes Layer - Semua endpoint API
Compatible dengan Supabase REST API
Termasuk 4 fitur baru: Review, Image Lightbox, Dashboard Stats, Tester Filter
"""
import os
import uuid
import re
from flask import request, jsonify, session
import openpyxl

# ============================================================
# IMPORTS - WAJIB ADA
# ============================================================
from database import query_db, execute_insert, execute_update, execute_delete

from models import (
    # Projects
    get_all_projects, get_project, create_project, update_project,
    archive_project, permanent_delete_project,
    # Scenarios
    get_scenarios_by_project, get_scenario, create_scenario,
    rename_scenario, update_scenario_meta, archive_scenario,
    hard_delete_scenario, update_scenario_status_logic,
    # Test Cases
    get_test_cases, add_test_case, update_test_case, delete_test_case,
    bulk_delete_test_cases, copy_test_case, reorder_test_cases,
    get_test_case_stats, get_last_tc_number, resequence_tcs_logic,
    # Attachments
    get_attachments, get_attachment_counts, save_screenshot, save_log,
    delete_screenshot, delete_log, rename_attachment, reorder_attachments,
    get_all_logs_for_export,
    # Search & Stats
    search_tickets, get_global_stats, get_project_ticket_counts,
    # FITUR BARU: Review & Period Stats
    toggle_review_status, get_review_stats,
    get_ticket_stats_by_period, get_tickets_by_tester,
)

from auth import (
    login_required, admin_required, reviewer_required,
    authenticate, login_user, logout_user, log_activity,
    verify_password, get_user_by_id, update_password,
    get_all_users, get_all_active_users,
    create_user_admin, update_user, delete_user,
    get_activity_logs,
)

from utils import call_ai_api


def register_api_routes(app):
    UPLOAD_FOLDER = app.config['UPLOAD_FOLDER']

    # ============================================================
    # AUTH ENDPOINTS
    # ============================================================
    @app.route('/api/auth/login', methods=['POST'])
    def api_login():
        data = request.get_json()
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        user, error = authenticate(username, password)
        if not user:
            return jsonify({'success': False, 'error': error or 'Login gagal'}), 401
        
        login_user(user)
        log_activity(user['id'], user['username'], 'login', ip=request.remote_addr)
        
        return jsonify({
            'success': True,
            'user': {
                'id': user['id'],
                'username': user['username'],
                'full_name': user.get('full_name', ''),
                'role': user.get('role', 'user')
            }
        })

    @app.route('/api/auth/logout', methods=['GET', 'POST'])
    @login_required
    def api_logout():
        log_activity(session.get('user_id'), session.get('username'),
                     'logout', ip=request.remote_addr)
        logout_user()
        session.clear()
        return jsonify({'success': True, 'message': 'Logout berhasil'})

    @app.route('/api/auth/change-password', methods=['POST'])
    @login_required
    def api_change_password():
        data = request.json
        old_password = data.get('old_password', '')
        new_password = data.get('new_password', '')
        
        if not old_password or not new_password:
            return jsonify({'error': 'Password wajib diisi'}), 400
        
        user_id = session.get('user_id')
        user = get_user_by_id(user_id)
        
        pwd_hash = user.get('password_hash')
        if not verify_password(old_password, pwd_hash):
            return jsonify({'error': 'Password lama salah'}), 401
        
        success, error = update_password(user_id, new_password)
        if not success:
            return jsonify({'error': error}), 400
        
        log_activity(user_id, user['username'], 'change_password',
                     target_type='user', target_id=user_id)
        return jsonify({'success': True, 'message': 'Password berhasil diubah'})

    # ============================================================
    # PROJECTS
    # ============================================================
    @app.route('/api/projects', methods=['GET'])
    @login_required
    def api_get_projects():
        return jsonify(get_all_projects())

    @app.route('/api/add_project', methods=['POST'])
    @login_required
    def api_add_project():
        name = request.json.get('name', '').strip()
        link = request.json.get('link', '')
        project, error = create_project(name, link)
        if not project:
            return jsonify({'error': error}), 400
        
        log_activity(session.get('user_id'), session.get('username'),
                     'create_project', target_type='project',
                     details=f'Created project: "{name}"')
        return jsonify({'success': True, 'project': project})

    @app.route('/api/update_project', methods=['POST'])
    @login_required
    def api_update_project():
        d = request.json
        update_project(d['id'], d.get('name', ''), d.get('link', ''))
        log_activity(session.get('user_id'), session.get('username'),
                     'update_project', target_type='project', target_id=d['id'])
        return jsonify({'success': True})

    @app.route('/api/archive_project/<int:pid>', methods=['POST'])
    @login_required
    def api_archive_project(pid):
        archive_project(pid, True)
        return jsonify({'success': True})

    @app.route('/api/restore_project/<int:pid>', methods=['POST'])
    @login_required
    def api_restore_project(pid):
        archive_project(pid, False)
        return jsonify({'success': True})

    @app.route('/api/permanent_delete_project/<int:pid>', methods=['POST'])
    @login_required
    def api_permanent_delete_project(pid):
        data = request.json
        proj = get_project(pid)
        if not proj or proj['name'] != data.get('confirm_name'):
            return jsonify({'error': 'Nama tidak sesuai'}), 403
        permanent_delete_project(pid)
        return jsonify({'success': True})

    # ============================================================
    # SCENARIOS (Tickets)
    # ============================================================
    @app.route('/api/scenarios/<int:pid>', methods=['GET'])
    @login_required
    def api_get_scenarios(pid):
        """Get all scenarios for project, with optional tester filter"""
        tester_filter = request.args.get('tester')
        scenarios = get_scenarios_by_project(pid, tester_filter=tester_filter)
        return jsonify(scenarios)

    @app.route('/api/scenario/<int:sid>', methods=['GET'])
    @login_required
    def api_get_scenario(sid):
        scenario = get_scenario(sid)
        if not scenario:
            return jsonify({'error': 'Scenario not found'}), 404
        
        project = get_project(scenario['project_id'])
        test_cases = get_test_cases(sid)
        stats = get_test_case_stats(sid)
        
        return jsonify({
            'scenario': scenario,
            'project': project,
            'test_cases': test_cases,
            'stats': stats
        })

    @app.route('/api/add_scenario/<int:pid>', methods=['POST'])
    @login_required
    def api_add_scenario(pid):
        if request.is_json:
            title = request.json.get('title', '').strip()
        else:
            title = request.form.get('title', '').strip()
        
        scenario, error = create_scenario(pid, title)
        if not scenario:
            return jsonify({'error': error}), 400
        
        log_activity(session.get('user_id'), session.get('username'),
                     'create_ticket', target_type='scenario',
                     target_id=scenario['id'],
                     details=f'Created ticket: "{title}" in Project #{pid}')
        return jsonify({'success': True, 'scenario': scenario})

    @app.route('/api/rename_scenario/<int:sid>', methods=['POST'])
    @login_required
    def api_rename_scenario(sid):
        title = request.json.get('title', '').strip()
        if not title:
            return jsonify({'error': 'Title required'}), 400
        rename_scenario(sid, title)
        log_activity(session.get('user_id'), session.get('username'),
                     'rename_ticket', target_type='scenario', target_id=sid,
                     details=f'Renamed to: "{title}"')
        return jsonify({'success': True})

    @app.route('/api/update_scenario_meta/<int:sid>', methods=['POST'])
    @login_required
    def api_update_scenario_meta(sid):
        d = request.json
        if d.get('end_date') and not d.get('start_date'):
            return jsonify({'error': 'Start Date harus diisi'}), 400
        if d.get('start_date') and d.get('end_date') and d['end_date'] < d['start_date']:
            return jsonify({'error': 'End Date tidak boleh sebelum Start Date'}), 400
        
        update_scenario_meta(sid, d.get('link', ''), d.get('testers', ''),
                             d.get('start_date', ''), d.get('end_date', ''))
        return jsonify({'success': True})

    @app.route('/api/archive_scenario/<int:sid>', methods=['POST'])
    @login_required
    def api_archive_scenario(sid):
        archive_scenario(sid, True)
        log_activity(session.get('user_id'), session.get('username'),
                     'archive_ticket', target_type='scenario', target_id=sid)
        return jsonify({'success': True})

    @app.route('/api/delete_scenario/<int:sid>', methods=['POST'])
    @login_required
    def api_delete_scenario_compat(sid):
        return api_archive_scenario(sid)

    @app.route('/api/restore_scenario/<int:sid>', methods=['POST'])
    @login_required
    def api_restore_scenario(sid):
        archive_scenario(sid, False)
        log_activity(session.get('user_id'), session.get('username'),
                     'restore_ticket', target_type='scenario', target_id=sid)
        return jsonify({'success': True})

    @app.route('/api/hard_delete_scenario/<int:sid>', methods=['POST'])
    @login_required
    def api_hard_delete_scenario(sid):
        hard_delete_scenario(sid)
        log_activity(session.get('user_id'), session.get('username'),
                     'hard_delete_ticket', target_type='scenario', target_id=sid)
        return jsonify({'success': True})

    # ============================================================
    # TEST CASES
    # ============================================================
    @app.route('/api/add_row/<int:sid>', methods=['POST'])
    @login_required
    def api_add_row(sid):
        tc_id, error = add_test_case(sid)
        if not tc_id:
            return jsonify({'error': error}), 500
        return jsonify({'tc_id': tc_id})

    @app.route('/api/update_cell/<int:sid>', methods=['POST'])
    @login_required
    def api_update_cell(sid):
        d = request.json
        success, error = update_test_case(sid, d['tc_id'], d['field'], d.get('value', ''))
        if not success:
            return jsonify({'error': error}), 400
        return jsonify({'success': True})

    @app.route('/api/delete_tc/<int:sid>', methods=['POST'])
    @login_required
    def api_delete_tc(sid):
        tc_id = request.json.get('tc_id')
        delete_test_case(sid, tc_id)
        log_activity(session.get('user_id'), session.get('username'),
                     'soft_delete_ticket', target_type='scenario', target_id=sid)
        return jsonify({'success': True})

    @app.route('/api/bulk_delete_tc/<int:sid>', methods=['POST'])
    @login_required
    def api_bulk_delete_tc(sid):
        tc_ids = request.json.get('tc_ids', [])
        if not tc_ids:
            return jsonify({'error': 'Tidak ada TC yang dipilih'}), 400
        
        deleted = bulk_delete_test_cases(sid, tc_ids)
        log_activity(session.get('user_id'), session.get('username'),
                     'bulk_delete_tc', target_type='scenario', target_id=sid,
                     details=f'Bulk deleted {deleted} TCs')
        return jsonify({'success': True, 'deleted': deleted})

    @app.route('/api/copy_tc/<int:sid>', methods=['POST'])
    @login_required
    def api_copy_tc(sid):
        source_tc_id = request.json.get('source_tc_id')
        new_tc_id, error = copy_test_case(sid, source_tc_id)
        if not new_tc_id:
            return jsonify({'error': error or 'Gagal copy'}), 400
        
        log_activity(session.get('user_id'), session.get('username'),
                     'copy_tc', target_type='test_case', target_id=new_tc_id,
                     details=f'Copied from {source_tc_id}')
        return jsonify({'success': True, 'new_tc_id': new_tc_id})

    @app.route('/api/reorder_tc/<int:sid>', methods=['POST'])
    @login_required
    def api_reorder_tc(sid):
        order = request.json.get('order', [])
        reorder_test_cases(sid, order)
        return jsonify({'success': True})

    @app.route('/api/summary/<int:sid>', methods=['GET'])
    def api_summary(sid):
        return jsonify(get_test_case_stats(sid))

    # ============================================================
    # REVIEW ENDPOINTS (FITUR BARU #1)
    # ============================================================
    @app.route('/api/toggle_review/<int:sid>', methods=['POST'])
    @reviewer_required
    def api_toggle_review(sid):
        """Toggle review status test case (hanya reviewer/admin)"""
        tc_id = request.json.get('tc_id')
        if not tc_id:
            return jsonify({'error': 'TC ID required'}), 400
        
        success, error = toggle_review_status(sid, tc_id, session.get('user_id'))
        if not success:
            return jsonify({'error': error}), 400
        
        log_activity(session.get('user_id'), session.get('username'),
                     'toggle_review', target_type='test_case', target_id=tc_id)
        return jsonify({'success': True})

    @app.route('/api/review_stats/<int:sid>', methods=['GET'])
    @login_required
    def api_review_stats(sid):
        """Get review statistics for scenario"""
        return jsonify(get_review_stats(sid))

    # ============================================================
    # ATTACHMENTS (Screenshots & Logs)
    # ============================================================
    @app.route('/api/get_attachments/<int:sid>/<tc_id>', methods=['GET'])
    @login_required
    def api_get_attachments(sid, tc_id):
        data = get_attachments(sid, tc_id)
        return jsonify({
            'screenshots': [{
                'id': s['id'],
                'path': s['file_path'],
                'url': f'/uploads/{s["file_path"]}',
                'name': s.get('custom_name') or os.path.basename(s['file_path'])
            } for s in data['screenshots']],
            'logs': [{
                'id': l['id'],
                'content': l.get('content', ''),
                'name': l.get('custom_name') or f"Log_{l['id']}"
            } for l in data['logs']]
        })

    @app.route('/api/attachment_counts/<int:sid>', methods=['GET'])
    @login_required
    def api_attachment_counts(sid):
        return jsonify(get_attachment_counts(sid))

    @app.route('/api/upload_screenshot/<int:sid>', methods=['POST'])
    @login_required
    def api_upload_screenshot(sid):
        """Upload screenshot untuk test case"""
        import uuid
        import os
        from datetime import datetime
        
        f = request.files.get('file')
        tc_id = request.form.get('tc_id')
        
        print(f"\n📸 Upload Request:")
        print(f"   - SID: {sid}")
        print(f"   - TC ID: {tc_id}")
        print(f"   - File: {f.filename if f else 'None'}")
        
        if not f or not tc_id:
            print(f"   ❌ Missing file or tc_id")
            return jsonify({'error': 'File atau TC ID tidak ditemukan'}), 400

        file_path = None
        try:
            # 1. Pastikan folder uploads ada
            upload_dir = app.config['UPLOAD_FOLDER']
            os.makedirs(upload_dir, exist_ok=True)
            print(f"   ✓ Upload dir: {upload_dir}")
            
            # 2. Buat nama file unik
            ext = os.path.splitext(f.filename)[1] if f.filename else '.png'
            name = f"{tc_id}_{uuid.uuid4().hex[:6]}{ext}"
            file_path = os.path.join(upload_dir, name)
            print(f"   ✓ File name: {name}")
            
            # 3. Simpan file
            f.save(file_path)
            print(f"   ✓ File saved to: {file_path}")
            
            # 4. VERIFIKASI: Cek apakah file benar-benar ada dan tidak kosong
            if not os.path.exists(file_path):
                raise Exception("File tidak ditemukan setelah save")
            
            file_size = os.path.getsize(file_path)
            print(f"   ✓ File size on disk: {file_size} bytes")
            
            if file_size == 0:
                raise Exception("File kosong (0 bytes)")

            # 5. Simpan ke Database - PASTIKAN file_path disimpan dengan benar
            from models import save_screenshot
            result = save_screenshot(sid, tc_id, name)  # Simpan nama file saja, bukan full path
            
            print(f"   ✓ Saved to DB. Result: {result}")
            
            # 6. Hitung total screenshot untuk TC ini
            from models import get_attachments
            attachments = get_attachments(sid, tc_id)
            count = len(attachments.get('screenshots', []))
            
            print(f"   ✓ Total screenshots: {count}")
            
            # 7. Log activity
            from auth import log_activity
            log_activity(
                session.get('user_id'), 
                session.get('username'), 
                'upload_screenshot', 
                target_type='test_case', 
                target_id=tc_id
            )
            
            print(f"   ✅ Upload successful!\n")
            
            # ✅ RETURN file_path yang benar (hanya nama file, bukan full path)
            return jsonify({
                'success': True, 
                'path': name,  # Hanya nama file
                'file_path': name,  # Tambahkan juga sebagai file_path
                'count': count
            })
            
        except Exception as e:
            print(f"   ❌ Upload failed: {str(e)}\n")
            import traceback
            traceback.print_exc()
            
            # Cleanup: Hapus file jika proses gagal
            if file_path and os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    print(f"   ✓ Cleaned up file: {file_path}")
                except:
                    pass
            return jsonify({'error': f'Upload gagal: {str(e)}'}), 500

    @app.route('/api/save_log/<int:sid>', methods=['POST'])
    @login_required
    def api_save_log(sid):
        d = request.json
        save_log(sid, d['tc_id'], d.get('content', ''), d.get('custom_name', ''))
        return jsonify({'success': True})

    @app.route('/api/delete_screenshot/<int:sid>', methods=['POST'])
    @login_required
    def api_delete_screenshot(sid):
        sc_id = request.json.get('id')
        if not sc_id:
            return jsonify({'error': 'ID required'}), 400
        delete_screenshot(sid, sc_id)
        return jsonify({'success': True})

    @app.route('/api/delete_log/<int:sid>', methods=['POST'])
    @login_required
    def api_delete_log(sid):
        log_id = request.json.get('id')
        if not log_id:
            return jsonify({'error': 'ID required'}), 400
        delete_log(sid, log_id)
        return jsonify({'success': True})

    @app.route('/api/rename_attachment/<int:sid>', methods=['POST'])
    @login_required
    def api_rename_attachment(sid):
        data = request.json
        rename_attachment(sid, data['id'], data.get('type', 'img'), data.get('name', ''))
        return jsonify({'success': True})

    @app.route('/api/reorder_attachments/<int:sid>', methods=['POST'])
    @login_required
    def api_reorder_attachments(sid):
        data = request.json
        reorder_attachments(sid, data.get('tc_id'), data.get('type', 'img'), data.get('order', []))
        return jsonify({'success': True})

    # ============================================================
    # AI & IMPORT
    # ============================================================
    @app.route('/api/ai_generate/<int:sid>', methods=['POST'])
    @login_required
    def api_ai_generate(sid):
        prompt = request.json.get('prompt', '')
        data = call_ai_api(prompt)
        
        tcs = get_test_cases(sid)
        max_ord = max([tc.get('display_order', 0) for tc in tcs], default=0)
        
        for i, item in enumerate(data, 1):
            execute_insert('test_cases', {
                'scenario_id': sid,
                'tc_id': f"TEMP-{i}",
                'display_order': max_ord + i,
                'test_case': item.get('test_case', ''),
                'test_criteria': item.get('test_criteria', ''),
                'test_date': item.get('test_date', ''),
                'test_data': item.get('test_data', ''),
                'expected_result': item.get('expected_result', ''),
                'actual_result': item.get('actual_result', ''),
                'status': 'Not Run',
                'remarks': item.get('remarks', ''),
                'is_deleted': False,
                'is_reviewed': False
            })
        
        resequence_tcs_logic(sid)
        update_scenario_status_logic(sid)
        
        return jsonify({'success': True, 'data': data})

    @app.route('/api/import_excel/<int:sid>', methods=['POST'])
    @login_required
    def import_excel_api(sid):
        """Import test cases dari Excel"""
        import traceback
        from datetime import datetime, date
        
        try:
            # 1. Validasi file
            if 'file' not in request.files:
                print("❌ No file in request")
                return jsonify({'error': 'Tidak ada file yang diupload'}), 400
            
            file = request.files['file']
            if not file.filename:
                return jsonify({'error': 'Nama file kosong'}), 400
            
            if not file.filename.lower().endswith('.xlsx'):
                return jsonify({'error': 'Hanya file .xlsx yang diperbolehkan'}), 400
            
            print(f"\n Import Excel Request:")
            print(f"   - SID: {sid}")
            print(f"   - File: {file.filename}")
            
            # 2. Load workbook
            try:
                wb = openpyxl.load_workbook(filename=file.stream, read_only=False, data_only=True)
            except Exception as e:
                print(f"   ❌ Gagal load workbook: {e}")
                return jsonify({'error': f'File Excel tidak valid: {str(e)}'}), 400
            
            ws = wb.active
            print(f"   ✓ Workbook loaded, sheet: {ws.title}, total rows: {ws.max_row}")
            
            # 3. Cari header - scan semua baris sampai row 20
            header_row = None
            header_row_idx = None
            
            for i in range(1, min(21, ws.max_row + 1)):
                row_values = [cell.value for cell in ws[i]]
                row_str = ' '.join([str(v).upper() if v else '' for v in row_values])
                
                # Cari baris yang mengandung "TC ID" atau "TEST CASE"
                if 'TC ID' in row_str or ('TEST' in row_str and 'CASE' in row_str):
                    header_row = [str(cell.value).strip().lower() if cell.value else '' 
                                for cell in ws[i]]
                    header_row_idx = i
                    print(f"   ✓ Header found at row {i}: {header_row}")
                    break
            
            if not header_row:
                print("   ❌ Header tidak ditemukan")
                return jsonify({'error': 'Header tidak ditemukan. Pastikan ada kolom "TC ID" atau "Test Case"'}), 400
            
            # 4. Mapping kolom
            mapping = {k: None for k in ['tc_id', 'test_case', 'test_criteria', 'test_date',
                                        'test_data', 'expected_result', 'actual_result', 
                                        'status', 'remarks']}
            
            keywords = {
                'tc_id': ['tc id', 'tc_id', 'tcid', 'id'],
                'test_case': ['test case', 'test_case', 'testcase', 'scenario', 'case'],
                'test_criteria': ['criteria', 'test criteria', 'test_criteria'],
                'test_date': ['test date', 'test_date', 'date', 'tanggal'],
                'test_data': ['data', 'test data', 'test_data', 'input'],
                'expected_result': ['expected', 'expected result', 'expected_result'],
                'actual_result': ['actual', 'actual result', 'actual_result'],
                'status': ['status', 'state'],
                'remarks': ['remarks', 'notes', 'log', 'keterangan', 'note']
            }
            
            for field, kws in keywords.items():
                for idx, h in enumerate(header_row):
                    if any(kw in h for kw in kws):
                        mapping[field] = idx
                        print(f"   ✓ Mapped {field} -> column {idx} ({h})")
                        break
            
            # 5. Helper function untuk konversi nilai
            def convert_excel_value(cell_value, field_name):
                if cell_value is None:
                    return ''
                if isinstance(cell_value, datetime):
                    if field_name == 'test_date':
                        return cell_value.strftime('%Y-%m-%d')
                    return cell_value.strftime('%Y-%m-%d %H:%M:%S')
                if isinstance(cell_value, date):
                    return cell_value.strftime('%Y-%m-%d')
                if isinstance(cell_value, (int, float)):
                    return str(int(cell_value)) if isinstance(cell_value, float) and cell_value.is_integer() else str(cell_value)
                return str(cell_value).strip()
            
            # 6. Proses data rows
            current_tc_num = get_last_tc_number(sid)
            tcs = get_test_cases(sid)
            max_ord = max([tc.get('display_order', 0) for tc in tcs], default=0)
            inserted = 0
            errors = []
            valid_statuses = {'Not Run', 'In Progress', 'Pass', 'Fail'}
            
            print(f"   ✓ Starting import from row {header_row_idx + 1} to {ws.max_row}")
            
            for row_idx in range(header_row_idx + 1, ws.max_row + 1):
                try:
                    row_values = [cell.value for cell in ws[row_idx]]
                    
                    # Skip baris kosong
                    if not any(cell for cell in row_values if cell is not None and str(cell).strip()):
                        continue
                    
                    def get_val(key):
                        idx = mapping[key]
                        if idx is not None and idx < len(row_values):
                            return convert_excel_value(row_values[idx], key)
                        return ''
                    
                    test_case = get_val('test_case')
                    if not test_case:
                        print(f"   ⚠️  Row {row_idx}: Skip - test case kosong")
                        continue
                    
                    status = get_val('status')
                    if status not in valid_statuses:
                        status = 'Not Run'
                    
                    tc_id = f"TC-{current_tc_num:03d}"
                    current_tc_num += 1
                    
                    # Prepare data
                    tc_data = {
                        'scenario_id': sid,
                        'tc_id': tc_id,
                        'display_order': max_ord + 1,
                        'test_case': test_case,
                        'test_criteria': get_val('test_criteria'),
                        'test_data': get_val('test_data'),
                        'expected_result': get_val('expected_result'),
                        'actual_result': get_val('actual_result'),
                        'status': status,
                        'remarks': get_val('remarks'),
                        'is_deleted': False,
                        'is_reviewed': False
                    }
                    
                    test_date = get_val('test_date')
                    if test_date:
                        tc_data['test_date'] = test_date
                    
                    # Insert ke database
                    execute_insert('test_cases', tc_data)
                    max_ord += 1
                    inserted += 1
                    
                except Exception as row_error:
                    error_msg = f"Row {row_idx}: {str(row_error)}"
                    errors.append(error_msg)
                    print(f"   ❌ Error di row {row_idx}: {row_error}")
                    print(f"   {traceback.format_exc()}")
                    continue
            
            # 7. Update status scenario
            update_scenario_status_logic(sid)
            
            # 8. Log activity
            log_activity(
                session.get('user_id'),
                session.get('username'),
                'import_excel',
                target_type='scenario',
                target_id=sid,
                details=f'Imported {inserted} test cases'
            )
            
            print(f"   ✅ Import selesai! Inserted: {inserted}, Errors: {len(errors)}\n")
            
            response_data = {
                'success': True,
                'imported': inserted,
                'errors': errors[:10] if errors else []
            }
            
            if errors:
                response_data['warning'] = f'{len(errors)} baris gagal diimport'
            
            return jsonify(response_data)
            
        except Exception as e:
            error_detail = str(e) + '\n' + traceback.format_exc()
            print(f"❌ Import Excel Error:\n{error_detail}")
            return jsonify({'error': f'Gagal import: {str(e)}'}), 500

    # ============================================================
    # SEARCH & STATISTICS
    # ============================================================
    @app.route('/api/search_tickets/<int:pid>', methods=['POST'])
    @login_required
    def api_search_tickets(pid):
        q = request.json.get('query', '').lower().strip()
        results = search_tickets(pid, q)
        return jsonify([{
            'id': r['id'],
            'title': r['title'],
            'status': r['status'],
            'deleted': bool(r.get('is_deleted'))
        } for r in results])

    @app.route('/api/global_stats', methods=['GET'])
    @login_required
    def api_global_stats():
        return jsonify(get_global_stats())

    @app.route('/api/project_ticket_counts', methods=['GET'])
    @login_required
    def api_project_ticket_counts():
        return jsonify(get_project_ticket_counts())

    # ============================================================
    # DASHBOARD STATS (FITUR BARU #3)
    # ============================================================
    @app.route('/api/ticket_stats_by_period', methods=['GET'])
    @login_required
    def api_ticket_stats_by_period():
        """Get ticket stats by month and week, with optional year/month filter"""
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        
        return jsonify(get_ticket_stats_by_period(year=year, month=month))
    
    @app.route('/api/available_months', methods=['GET'])
    @login_required
    def api_available_months():
        """Get list of months that have tickets (for filter dropdown)"""
        scenarios = query_db('scenarios', 
                            filters={'is_deleted': False},
                            order='created_at.desc')
        
        months_set = set()
        for s in scenarios:
            created_at = s.get('created_at', '')
            if not created_at:
                continue
            
            try:
                if 'T' in created_at:
                    date_str = created_at.split('T')[0]
                else:
                    date_str = created_at.split(' ')[0]
                
                year_month = date_str[:7]  # 2026-06
                months_set.add(year_month)
            except:
                continue
        
        # Convert ke list dan sort descending
        months_list = sorted(list(months_set), reverse=True)
        
        result = []
        for ym in months_list:
            year, month = ym.split('-')
            month_names = [
                'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember'
            ]
            result.append({
                'value': ym,
                'label': f"{month_names[int(month)-1]} {year}",
                'year': int(year),
                'month': int(month)
            })
        
        return jsonify(result)

    @app.route('/api/tickets_by_tester', methods=['GET'])
    @login_required
    def api_tickets_by_tester():
        """Get ticket counts grouped by tester"""
        return jsonify(get_tickets_by_tester())

    # ============================================================
    # USER LIST (FITUR BARU #4 - untuk dropdown tester)
    # ============================================================
    @app.route('/api/users_list', methods=['GET'])
    @login_required
    def api_users_list():
        """Get list of active users (untuk dropdown tester)"""
        return jsonify(get_all_active_users())

    # ============================================================
    # ADMIN ENDPOINTS
    # ============================================================
    @app.route('/api/admin/users', methods=['GET'])
    @admin_required
    def api_admin_users():
        return jsonify(get_all_users())

    @app.route('/api/admin/users', methods=['POST'])
    @admin_required
    def api_admin_create_user():
        data = request.json
        user_id, error = create_user_admin(
            data.get('username', ''),
            data.get('password', ''),
            data.get('full_name', ''),
            data.get('role', 'user'),
            data.get('email', '')
        )
        if not user_id:
            return jsonify({'error': error}), 400
        
        log_activity(session.get('user_id'), session.get('username'),
                     'create_user', target_type='user', target_id=user_id,
                     details=f'Created user: {data.get("username")}')
        return jsonify({'success': True, 'user_id': user_id})

    @app.route('/api/admin/users/<int:user_id>', methods=['PUT'])
    @admin_required
    def api_admin_update_user(user_id):
        data = request.json
        success, error = update_user(
            user_id,
            full_name=data.get('full_name', ''),
            email=data.get('email', ''),
            role=data.get('role', 'user'),
            is_active=data.get('is_active', True),
            new_password=data.get('new_password') or None
        )
        if not success:
            return jsonify({'error': error}), 400
        
        log_activity(session.get('user_id'), session.get('username'),
                     'update_user', target_type='user', target_id=user_id)
        return jsonify({'success': True})

    @app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
    @admin_required
    def api_admin_delete_user(user_id):
        success, error = delete_user(user_id)
        if not success:
            return jsonify({'error': error}), 403
        
        log_activity(session.get('user_id'), session.get('username'),
                     'delete_user', target_type='user', target_id=user_id)
        return jsonify({'success': True})

    @app.route('/api/admin/activity-logs', methods=['GET'])
    @login_required
    def api_admin_activity_logs():
        page = request.args.get('page', 1, type=int)
        limit = request.args.get('limit', 50, type=int)
        user_id = request.args.get('user_id', type=int)
        action = request.args.get('action')
        
        logs, total = get_activity_logs(
            limit=limit, user_id=user_id, action=action, page=page
        )
        
        return jsonify({
            'logs': logs,
            'total': total,
            'page': page,
            'limit': limit,
            'total_pages': (total + limit - 1) // limit if limit > 0 else 0
        })