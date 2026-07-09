"""
Web Routes - Halaman HTML (Jinja Templates)
Compatible dengan Supabase REST API
"""
from flask import render_template, redirect, url_for, flash, session
from auth import login_required, admin_required, get_current_user
from models import (
    get_all_projects, get_project, get_scenarios_by_project,
    get_scenario, get_test_cases, get_test_case_stats, get_all_logs_for_export,
    get_attachment_counts,
)
from utils import generate_export_filename
from services import get_dashboard_analytics, get_project_dashboard_data, prepare_scenario_detail


def register_web_routes(app):
    
    @app.route('/')
    @login_required
    def index():
        """Dashboard utama"""
        analytics = get_dashboard_analytics()
        projects = get_all_projects()
        return render_template('dashboard.html', 
                             analytics=analytics,
                             projects=projects,
                             user=get_current_user())
    
    @app.route('/project/<int:pid>')
    @login_required
    def view_project(pid):
        """Halaman project detail"""
        data = get_project_dashboard_data(pid)
        if not data:
            flash('Project tidak ditemukan', 'danger')
            return redirect(url_for('index'))
        
        return render_template('project.html', **data, user=get_current_user())
    
    @app.route('/scenario/<int:sid>')
    @login_required
    def view_scenario(sid):
        """Halaman scenario detail"""
        data = prepare_scenario_detail(sid)
        if not data:
            flash('Scenario tidak ditemukan', 'danger')
            return redirect(url_for('index'))
        
        return render_template('scenario.html', **data, user=get_current_user())
    
    @app.route('/export/<int:sid>')
    @login_required
    def export_excel(sid):
        """Export scenario ke Excel"""
        import io
        import re
        from flask import send_file
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from database import query_db
        
        # ============================================
        # ✅ FUNGSI SANITASI: Hapus Karakter Ilegal XML/Excel
        # ============================================
        # Karakter ilegal untuk XML: \x00-\x08, \x0B-\x0C, \x0E-\x1F, \x7F
        ILLEGAL_CHARS_REGEX = re.compile(
            r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]'
        )
        
        def sanitize_value(value):
            """Bersihkan karakter ilegal dari string"""
            if value is None:
                return ''
            if not isinstance(value, str):
                value = str(value)
            # Hapus karakter ilegal
            return ILLEGAL_CHARS_REGEX.sub('', value)
        
        # 1. Ambil Data
        scenario = query_db('scenarios', filters={'id': sid}, fetch='one')
        if not scenario:
            return "Scenario not found in database", 404
            
        project = query_db('projects', filters={'id': scenario['project_id']}, fetch='one')
        project_name = project['name'] if project else 'Unknown Project'
        
        test_cases = get_test_cases(sid, include_deleted=False)
        logs = get_all_logs_for_export(sid)
        att_counts = get_attachment_counts(sid)
        
        # 2. Buat Workbook
        wb = Workbook()
        
        # --- Sheet 1: Test Cases ---
        ws_tc = wb.active
        ws_tc.title = "Test Cases"
        
        headers = ['TC ID', 'Test Case', 'Criteria', 'Test Date', 'Test Data', 
                'Expected Result', 'Actual Result', 'Status', 'Remarks', 'Imgs', 'Logs']
        ws_tc.append(headers)
        
        # Styling Header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws_tc[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
            
        # Masukkan Data Test Cases
        for tc in test_cases:
            tc_id = tc.get('tc_id', '')
            counts = att_counts.get(tc_id, {'img': 0, 'log': 0})
            
            # ✅ SANITASI semua field sebelum ditulis ke Excel
            ws_tc.append([
                sanitize_value(tc_id),
                sanitize_value(tc.get('test_case', '')),
                sanitize_value(tc.get('test_criteria', '')),
                sanitize_value(tc.get('test_date', '')),
                sanitize_value(tc.get('test_data', '')),
                sanitize_value(tc.get('expected_result', '')),
                sanitize_value(tc.get('actual_result', '')),
                sanitize_value(tc.get('status', '')),
                sanitize_value(tc.get('remarks', '')),
                counts['img'],
                counts['log']
            ])
            
        # Auto-width kolom (basic)
        for col in ws_tc.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_tc.column_dimensions[column].width = min(max_length + 2, 50)

        # --- Sheet 2: Logs ---
        ws_logs = wb.create_sheet(title="Logs")
        ws_logs.append(['TC ID', 'Log Name', 'Content'])
        
        for cell in ws_logs[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            
        for log in logs:
            # ✅ SANITASI content log (ini biasanya sumber utama karakter ilegal)
            ws_logs.append([
                sanitize_value(log.get('tc_id', '')),
                sanitize_value(log.get('custom_name') or f"Log_{log.get('id')}"),
                sanitize_value(log.get('content', ''))  # ⚠️ Ini yang paling sering bermasalah
            ])
            
        # Auto-width kolom Logs
        for col in ws_logs.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_logs.column_dimensions[column].width = min(max_length + 2, 80)

        # 3. Simpan ke Memory (BytesIO)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = generate_export_filename(scenario, project)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )