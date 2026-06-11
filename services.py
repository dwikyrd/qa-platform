"""
Services Layer - Business Logic untuk Export & Dashboard
Compatible dengan Supabase REST API
"""
import os
import io
import html
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

try:
    from openpyxl.drawing.image import Image as OpenPyxlImage
    HAS_PILLOW = True
except ImportError:
    HAS_PILLOW = False

# ✅ Import yang BENAR dari models
from models import (
    get_project, get_scenarios_by_project, get_scenario, get_test_cases,
    get_attachments, get_all_logs_for_export, get_test_case_stats,
    get_global_stats, get_project_ticket_counts
)
from utils import calc_duration, generate_export_filename

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')


# ================= DASHBOARD SERVICES =================
def get_dashboard_analytics():
    """Menyiapkan data visualisasi chart berdasarkan TICKET"""
    stats = get_global_stats()
    
    chart_status = {
        'Done': stats['tickets'].get('done', 0),
        'In Progress': stats['tickets'].get('in_progress', 0),
        'Not Run': stats['tickets'].get('not_run', 0),
        'Fail': stats['tickets'].get('fail', 0)
    }
    
    project_counts = get_project_ticket_counts()
    
    return {
        'status_chart': chart_status,
        'project_chart': project_counts
    }


def get_project_dashboard_data(pid):
    """Data untuk halaman project"""
    proj = get_project(pid)
    if not proj:
        return None
    
    all_scenarios = get_scenarios_by_project(pid)
    active_scenarios = [s for s in all_scenarios if not s.get('is_deleted')]
    archived_scenarios = [s for s in all_scenarios if s.get('is_deleted')]
    
    stats = {
        'total': len(active_scenarios),
        'active': len([s for s in active_scenarios if s.get('status') == 'In Progress']),
        'inactive': len([s for s in active_scenarios if s.get('status') == 'Not Run']),
        'completed': len([s for s in active_scenarios if s.get('status') == 'Done'])
    }
    
    # Timeline data
    recent_data = [{
        'id': r['id'], 
        'title': r['title'], 
        'status': r.get('status'),
        'duration': calc_duration(r.get('start_date'), r.get('end_date'), r.get('status')),
        'created': r.get('created_at', '')[:10] if r.get('created_at') else '-'
    } for r in active_scenarios[:5]]
    
    timeline_data = [{
        'id': t['id'], 
        'title': t['title'], 
        'status': t.get('status'),
        'start': t.get('start_date') or '-', 
        'end': t.get('end_date') or '-',
        'duration': calc_duration(t.get('start_date'), t.get('end_date'), t.get('status'))
    } for t in active_scenarios]
    
    return {
        'project': proj, 
        'scenarios': active_scenarios, 
        'archived': archived_scenarios,
        'pid': pid, 
        'stats': stats, 
        'recent': recent_data, 
        'timeline': timeline_data
    }


def prepare_scenario_detail(sid):
    """Data untuk halaman scenario detail"""
    sc = get_scenario(sid)
    if not sc:
        return None
    
    proj = get_project(sc['project_id'])
    tcs = get_test_cases(sid)
    
    # Escape HTML untuk keamanan
    for tc in tcs:
        for field in ['test_case', 'test_criteria', 'test_data',
                      'expected_result', 'actual_result', 'remarks']:
            if tc.get(field):
                tc[field] = html.escape(str(tc[field]), quote=False)
    
    return {
        'scenario': sc,
        'project': proj,
        'test_cases': tcs,
        'stats': get_test_case_stats(sid)
    }


# ================= EXPORT EXCEL SERVICE =================
def export_to_excel(sid):
    """Export test cases ke Excel dengan handling screenshot yang lebih baik"""
    import os
    import io
    import html
    import traceback
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    try:
        from openpyxl.drawing.image import Image as OpenPyxlImage
        HAS_PILLOW = True
    except ImportError:
        HAS_PILLOW = False
        print("⚠️  Pillow/OpenPyxl Image not available. Screenshots will be skipped.")
    
    # Import models
    from models import (
        get_project, get_scenario, get_test_cases,
        get_attachments, get_all_logs_for_export
    )
    from utils import generate_export_filename
    
    # ===== PENTING: Tentukan UPLOAD_FOLDER dengan benar =====
    # Gunakan path absolut dari lokasi file ini (services.py)
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
    
    print(f"\n📊 Starting Excel Export for SID: {sid}")
    print(f"📁 Upload folder: {UPLOAD_FOLDER}")
    print(f"📁 Upload folder exists: {os.path.exists(UPLOAD_FOLDER)}")
    
    try:
        # 1. Ambil data scenario & project
        sc = get_scenario(sid)
        if not sc:
            return None, "Scenario not found"
        
        proj = get_project(sc['project_id'])
        tcs = get_test_cases(sid)
        atts = get_attachments(sid, '')
        all_logs = get_all_logs_for_export(sid)
        
        print(f"📋 Test Cases: {len(tcs)}")
        print(f"📎 Screenshots: {len(atts.get('screenshots', []))}")
        print(f"📝 Logs: {len(all_logs)}")
        
        # 2. Group screenshots by tc_id
        sc_dict = {}
        for r in atts.get('screenshots', []):
            tc_id = r.get('tc_id')
            if tc_id:
                if tc_id not in sc_dict:
                    sc_dict[tc_id] = []
                sc_dict[tc_id].append({
                    'path': r.get('file_path', ''),
                    'name': r.get('custom_name') or os.path.basename(r.get('file_path', ''))
                })
        
        print(f"📊 Screenshots by TC: {len(sc_dict)} test cases have screenshots")
        
        # 3. Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Execution Report"
        
        # Header
        header = ws.cell(row=1, column=1, value="System Integration Testing")
        header.font = Font(size=20, bold=True, color="1F4E79")
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=2)
        
        # Metadata
        full_link = sc.get('link') or (proj.get('link') if proj else '-')
        meta = [
            ("Project Name", proj.get('name') if proj else 'N/A'),
            ("Project Link", full_link),
            ("Testing Start Date / Time", sc.get('start_date') or "-"),
            ("Testing End Date / Time", sc.get('end_date') or "-"),
            ("Name of Tester/s:", sc.get('testers') or "-")
        ]
        
        for i, (label, value) in enumerate(meta, 2):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)
        
        # Test cases header
        hdrs = ["TC ID", "Test Case", "Test Criteria", "Test Date", "Test Data",
                "Expected Result", "Actual Result", "Status", "Remarks"]
        start_row = len(meta) + 2
        
        for c, h in enumerate(hdrs, 1):
            cell = ws.cell(row=start_row, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center")
        
        # Test cases data
        for r, tc in enumerate(tcs, start_row + 1):
            ws.cell(row=r, column=1, value=tc.get('tc_id'))
            ws.cell(row=r, column=2, value=tc.get('test_case') or "-")
            
            for c, k in enumerate(['test_criteria', 'test_date', 'test_data',
                                   'expected_result', 'actual_result', 'status', 'remarks'], 3):
                val = tc.get(k)
                if val and isinstance(val, str):
                    val = val.strip()
                    cell = ws.cell(row=r, column=c, value=val if val else "-")
                    if '\n' in val or len(val) > 50:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                else:
                    ws.cell(row=r, column=c, value="-")
                
                if k == 'status':
                    col = "008000" if val == "Pass" else "FF0000" if val == "Fail" \
                          else "FFA500" if val == "In Progress" else "808080"
                    ws.cell(row=r, column=c).font = Font(bold=True, color=col)
            
            # Border
            for c in range(1, 10):
                ws.cell(row=r, column=c).border = Border(
                    bottom=Side(style="thin"), top=Side(style="thin"),
                    left=Side(style="thin"), right=Side(style="thin")
                )
            
            ws.row_dimensions[r].height = 20
        
        # ===== SHEET SCREENSHOTS (FIXED) =====
        ws2 = wb.create_sheet("Screenshots")
        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 35
        ws2.column_dimensions['C'].width = 45
        
        for i, h in enumerate(["TC ID", "Photo Name", "Preview"], 1):
            ws2.cell(row=1, column=i, value=h).font = Font(bold=True)
        
        ri = 2
        images_added = 0
        images_failed = 0
        
        for tc in tcs:
            tc_id = tc.get('tc_id')
            screenshots = sc_dict.get(tc_id, [])
            
            for img in screenshots:
                ws2.cell(row=ri, column=1, value=tc_id)
                ws2.cell(row=ri, column=2, value=img['name'])
                
                # Build full path
                img_filename = img['path']
                full_path = os.path.join(UPLOAD_FOLDER, img_filename)
                
                print(f"   🖼️  Processing: {img_filename}")
                print(f"      Full path: {full_path}")
                print(f"      Exists: {os.path.exists(full_path)}")
                
                try:
                    # Cek apakah file ada dan tidak kosong
                    if not os.path.exists(full_path):
                        raise FileNotFoundError(f"File not found: {full_path}")
                    
                    file_size = os.path.getsize(full_path)
                    if file_size == 0:
                        raise ValueError(f"File is empty (0 bytes): {full_path}")
                    
                    print(f"      Size: {file_size} bytes")
                    
                    # Cek apakah Pillow/OpenPyxl Image tersedia
                    if not HAS_PILLOW:
                        ws2.cell(row=ri, column=3, 
                                value=f"⚠️ Pillow not installed - {img_filename}")
                        images_failed += 1
                    else:
                        # Coba load image
                        im = OpenPyxlImage(full_path)
                        
                        # Resize jika terlalu besar
                        if im.width > 300:
                            ratio = 300 / im.width
                            im.width = 300
                            im.height = int(im.height * ratio)
                        
                        # Tambahkan ke cell
                        cell_ref = f"C{ri}"
                        ws2.add_image(im, cell_ref)
                        
                        # Set row height sesuai tinggi gambar
                        ws2.row_dimensions[ri].height = (im.height * 0.75) + 15
                        
                        images_added += 1
                        print(f"      ✅ Image added successfully")
                        
                except Exception as img_error:
                    error_msg = f"❌ Error: {str(img_error)[:100]}"
                    ws2.cell(row=ri, column=3, value=error_msg)
                    images_failed += 1
                    print(f"      {error_msg}")
                    # Lanjut ke gambar berikutnya, jangan hentikan export
                
                ri += 1
        
        print(f"\n📊 Screenshot Summary:")
        print(f"   ✅ Added: {images_added}")
        print(f"   ❌ Failed: {images_failed}")
        
        # ===== SHEET LOG DATA =====
        ws3 = wb.create_sheet("Log Data")
        for i, h in enumerate(["TC ID", "Log Name", "Content"], 1):
            ws3.cell(row=1, column=i, value=h).font = Font(bold=True)
        
        for i, log in enumerate(all_logs, 2):
            ws3.cell(row=i, column=1, value=log.get('tc_id'))
            ws3.cell(row=i, column=2, value=log.get('custom_name') or f"Log_{log.get('id', i)}")
            ws3.cell(row=i, column=3, value=log.get('content', ''))
            ws3.cell(row=i, column=3).alignment = Alignment(wrap_text=True)
        
        # Save ke buffer
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        filename = generate_export_filename(sc, proj)
        print(f"\n✅ Export completed: {filename}")
        print(f"   Total screenshots: {images_added} added, {images_failed} failed\n")
        
        return buf, filename
        
    except Exception as e:
        error_detail = traceback.format_exc()
        print(f"\n❌ Export error: {error_detail}")
        return None, str(e)
    

    